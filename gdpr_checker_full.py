import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import re
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# -----------------------------
# Configuration
# -----------------------------
START_URL = "https://webdriveruniversity.com/index.html"  # Replace with your dummy/test site
MAX_PAGES = 10
TIMEOUT = 30
RETRIES = 3
RETRY_DELAY = 5

# Regex patterns for PII
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'\+?\d[\d\-\s]{7,}\d')

# Weighted checklist for compliance scoring
CHECKLIST = {
    'Privacy policy link': 2,
    'Cookie consent mechanism': 2,
    'HTTPS enforced': 2,
    'Forms ask unnecessary fields': 1,
    'Email detected': 1,
    'Phone detected': 1,
    'Third-party trackers': 1
}

# -----------------------------
# Helper Functions
# -----------------------------
def fetch_page(url):
    for attempt in range(RETRIES):
        try:
            r = requests.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            return r.text, r.url
        except requests.RequestException as e:
            print(f"Attempt {attempt+1} failed for {url}: {e}")
            if attempt < RETRIES - 1:
                time.sleep(RETRY_DELAY)
            else:
                print(f"Skipping {url} after {RETRIES} failed attempts.")
    return None, url

def parse_forms(html, base_url):
    soup = BeautifulSoup(html, 'html.parser')
    forms = []
    for f in soup.find_all('form'):
        action = f.get('action') or ''
        action = urljoin(base_url, action)
        inputs = []
        for inp in f.find_all(['input','textarea','select']):
            inputs.append({
                'name': inp.get('name'),
                'type': inp.get('type'),
                'placeholder': inp.get('placeholder'),
                'autocomplete': inp.get('autocomplete')
            })
        forms.append({'action': action, 'inputs': inputs})
    return forms

def find_pii(html):
    emails = set(EMAIL_RE.findall(html))
    phones = set(PHONE_RE.findall(html))
    return {'emails': list(emails), 'phones': list(phones)}

def detect_cookies(html):
    soup = BeautifulSoup(html, 'html.parser')
    text = soup.get_text().lower()
    cookie_keywords = ['cookie', 'consent', 'privacy', 'tracking']
    found = [kw for kw in cookie_keywords if kw in text]
    return found

def detect_https(url):
    return url.startswith("https://")

def find_links(html, base_url):
    soup = BeautifulSoup(html, 'html.parser')
    links = []
    for a in soup.find_all('a', href=True):
        href = urljoin(base_url, a['href'])
        if urlparse(href).netloc == urlparse(base_url).netloc:
            links.append(href)
    return list(set(links))

def add_recommendation(issue_type, value=None):
    if issue_type == 'Email detected':
        return ("Verify storage and lawful basis for email collection",
                "GDPR Art. 5 & 6", "High")
    elif issue_type == 'Phone detected':
        return ("Verify storage and lawful basis for phone numbers",
                "GDPR Art. 5 & 6", "High")
    elif 'Form field' in issue_type:
        return ("Check if form collects only necessary data, obtain consent if needed",
                "GDPR Art. 5 & 6", "Medium")
    elif 'Cookie' in issue_type:
        return ("Add explicit consent mechanism for cookies/tracking",
                "GDPR Art. 6 & 7", "High")
    elif issue_type == 'HTTPS enforced':
        return ("Enforce HTTPS on all pages to secure data in transit",
                "GDPR Art. 32", "High")
    else:
        return ("Review issue manually", "GDPR general", "Low")

# -----------------------------
# Main Scanner
# -----------------------------
visited = set()
to_visit = [START_URL]
report_rows = []

while to_visit and len(visited) < MAX_PAGES:
    url = to_visit.pop(0)
    if url in visited:
        continue

    html, final_url = fetch_page(url)
    if html is None:
        continue

    visited.add(final_url)

    # Forms
    forms = parse_forms(html, final_url)
    for f in forms:
        for inp in f['inputs']:
            issue = f"Form field: {inp['name']}"
            rec, legal, priority = add_recommendation(issue)
            report_rows.append({
                'page': final_url,
                'issue': issue,
                'value': inp['name'],
                'result': 'Detected',
                'recommendation': rec,
                'legal_reference': legal,
                'priority': priority
            })

    # PII
    pii = find_pii(html)
    for email in pii['emails']:
        rec, legal, priority = add_recommendation('Email detected')
        report_rows.append({
            'page': final_url,
            'issue': 'Email detected',
            'value': email,
            'result': 'Flagged',
            'recommendation': rec,
            'legal_reference': legal,
            'priority': priority
        })
    for phone in pii['phones']:
        rec, legal, priority = add_recommendation('Phone detected')
        report_rows.append({
            'page': final_url,
            'issue': 'Phone detected',
            'value': phone,
            'result': 'Flagged',
            'recommendation': rec,
            'legal_reference': legal,
            'priority': priority
        })

    # Cookies
    cookies = detect_cookies(html)
    for kw in cookies:
        rec, legal, priority = add_recommendation('Cookie/consent keyword')
        report_rows.append({
            'page': final_url,
            'issue': f'Cookie/consent keyword: {kw}',
            'value': kw,
            'result': 'Detected',
            'recommendation': rec,
            'legal_reference': legal,
            'priority': priority
        })

    # HTTPS check
    if not detect_https(final_url):
        rec, legal, priority = add_recommendation('HTTPS enforced')
        report_rows.append({
            'page': final_url,
            'issue': 'HTTPS enforced',
            'value': '',
            'result': 'Not compliant',
            'recommendation': rec,
            'legal_reference': legal,
            'priority': priority
        })

    # Internal links
    links = find_links(html, final_url)
    for link in links:
        if link not in visited and link not in to_visit:
            to_visit.append(link)

# -----------------------------
# Compliance Scoring
# -----------------------------
df = pd.DataFrame(report_rows)
score_total = 0
weight_total = 0
for check, weight in CHECKLIST.items():
    occurrences = df[df['issue'].str.contains(check, case=False)]
    if not occurrences.empty:
        score_total += weight
    weight_total += weight

compliance_percent = (score_total / weight_total) * 100 if weight_total else 0

# -----------------------------
# Save CSV report
# -----------------------------
csv_file = 'gdpr_report_full.csv'
df.to_csv(csv_file, index=False)

# -----------------------------
# Create Excel Summary
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "GDPR Report"

# Header
headers = ['Page', 'Issue', 'Value', 'Result', 'Recommendation', 'Legal Reference', 'Priority']
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Fill data
for index, row in df.iterrows():
    ws.append([row['page'], row['issue'], row['value'], row['result'],
               row['recommendation'], row['legal_reference'], row['priority']])

# Executive Summary Sheet
summary = wb.create_sheet(title="Summary")
summary.append(["GDPR Compliance Summary"])
summary.append(["Total pages scanned:", len(visited)])
summary.append(["Compliance Score (%):", f"{compliance_percent:.2f}%"])
summary.append(["Top 5 issues:"])
top_issues = df['issue'].value_counts().head(5)
for issue, count in top_issues.items():
    summary.append([issue, count])

excel_file = 'gdpr_report_summary.xlsx'
wb.save(excel_file)

# -----------------------------
# Print Summary
# -----------------------------
print("GDPR Compliance Scanner finished!")
print(f"Pages successfully scanned: {len(visited)}")
print(f"Compliance Score: {compliance_percent:.2f}%")
print(f"CSV report saved as: {csv_file}")
print(f"Excel summary report saved as: {excel_file}")
print("Top 5 detected issues:")
for issue, count in top_issues.items():
    print(f"- {issue}: {count}")
